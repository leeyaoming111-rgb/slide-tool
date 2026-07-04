#!/usr/bin/env python3
"""Generate templates/chart-templates.xlsx — native Excel chart templates styled to the
case-design-system tokens (tokens/tokens.json) and house rules (docs/CHART_RULES.md).

One tab per chart type. Each tab has a clearly marked INPUT block (blue font = type here),
grey helper formulas where the chart needs scaffolding (waterfall bases etc.), and a live
chart wired to the inputs. Corporate family only.

Run:  python3 templates/build_chart_templates.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, RadarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint, Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    Paragraph, ParagraphProperties, CharacterProperties, Font as DrawFont,
)

# ---- tokens (concrete hex, mirrors tokens/tokens.json) -------------------------------
INK, NAVY, NAVYDEEP = "14233A", "1F3A52", "163041"
ACCENT, BLUEMID, BLUELIGHT = "2E6FB0", "5B8FC7", "A9C4DD"
SLATE, GRID, PANEL, WHITE = "5B6B79", "C9D2DA", "F4F6F7", "FFFFFF"
POS, POSBG, NEG, NEGBG, CAUT, CAUTBG = "2E8B6F", "C9E7D6", "C0473E", "F3D8D5", "E0A13C", "F7E8CC"

FONT = "Inter"  # falls back to Arial when Inter is not installed

# ---- cell style shorthand ------------------------------------------------------------
F_TITLE  = Font(name=FONT, size=14, bold=True, color=INK)
F_KICKER = Font(name=FONT, size=9,  bold=True, color=NAVY)
F_BODY   = Font(name=FONT, size=10, color=INK)
F_BOLD   = Font(name=FONT, size=10, bold=True, color=INK)
F_MUTED  = Font(name=FONT, size=9,  color=SLATE)
F_SOURCE = Font(name=FONT, size=9,  italic=True, color=SLATE)
F_INPUT  = Font(name=FONT, size=10, color=ACCENT)          # blue font = hardcoded input
F_CALC   = Font(name=FONT, size=10, color=SLATE)            # grey font = helper formula
F_HDRW   = Font(name=FONT, size=10, bold=True, color=WHITE)

FILL_NAVY  = PatternFill("solid", fgColor=NAVY)
FILL_PANEL = PatternFill("solid", fgColor=PANEL)
FILL_POSBG = PatternFill("solid", fgColor=POSBG)
FILL_CAUTBG= PatternFill("solid", fgColor=CAUTBG)

RIGHT  = Alignment(horizontal="right")
CENTER = Alignment(horizontal="center")
HAIR   = Side(style="thin", color=GRID)
RULE_B = Border(bottom=Side(style="medium", color=NAVY))
RULE_T = Border(top=Side(style="thin", color=NAVY))
BOX    = Border(top=HAIR, bottom=HAIR, left=HAIR, right=HAIR)

NUM   = "#,##0"
NUM1  = "#,##0.0"
PCT   = "0%"
PCT1  = "0.0%"
USD   = '"$"#,##0'


# ---- chart style helpers -------------------------------------------------------------
def rich(sz=1100, color=SLATE, b=False):
    cp = CharacterProperties(latin=DrawFont(typeface=FONT), sz=sz, solidFill=color, b=b)
    return RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])


def faint_gridlines():
    return ChartLines(spPr=GraphicalProperties(ln=LineProperties(solidFill=GRID, w=9525)))


def style_axes(ch, cat_font=True, val_delete=False, val_grid=False, val_fmt=None):
    """House axis treatment: thin grey category baseline, no top/right clutter,
    value axis deleted when data labels carry the numbers (rule 2)."""
    ch.x_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=GRID, w=9525))
    ch.x_axis.majorGridlines = None
    ch.x_axis.majorTickMark = "none"
    if cat_font:
        ch.x_axis.txPr = rich(1100, SLATE)
    ch.y_axis.majorTickMark = "none"
    ch.y_axis.majorGridlines = faint_gridlines() if val_grid else None
    if val_delete:
        ch.y_axis.delete = True
    else:
        ch.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
        ch.y_axis.txPr = rich(1100, SLATE)
        if val_fmt:
            ch.y_axis.number_format = val_fmt
    ch.roundedCorners = False
    ch.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))


def solid(series, hexcolor, line=False):
    gp = GraphicalProperties(solidFill=hexcolor)
    gp.ln = LineProperties(solidFill=hexcolor, w=9525) if line else LineProperties(noFill=True)
    series.spPr = gp


def invisible(series):
    gp = GraphicalProperties()
    gp.noFill = True
    gp.ln = LineProperties(noFill=True)
    series.spPr = gp


def line_style(series, hexcolor, w_pt=2.25):
    gp = GraphicalProperties()
    gp.ln = LineProperties(solidFill=hexcolor, w=int(w_pt * 12700))
    series.spPr = gp
    series.marker = Marker(symbol="none")
    series.smooth = False


def labels(series_or_chart, fmt=NUM, pos=None, color=INK, sz=1000, b=True):
    dl = DataLabelList(showVal=True, showSerName=False, showCatName=False,
                       showLegendKey=False, showPercent=False, showBubbleSize=False)
    dl.numFmt = fmt
    if pos:
        dl.dLblPos = pos
    dl.txPr = rich(sz, color, b)
    series_or_chart.dLbls = dl
    return dl


def point_color(series, idx, hexcolor):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties = GraphicalProperties(solidFill=hexcolor)
    pt.graphicalProperties.ln = LineProperties(noFill=True)
    series.dPt.append(pt)


# ---- sheet scaffolding ---------------------------------------------------------------
def sheet(wb, name, kicker, title, use_note, tab=NAVY):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab
    ws.sheet_view.showGridLines = False
    ws["A1"] = kicker.upper()
    ws["A1"].font = F_KICKER
    ws["A2"] = title
    ws["A2"].font = F_TITLE
    ws["A3"] = use_note
    ws["A3"].font = F_MUTED
    ws.column_dimensions["A"].width = 26
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 11
    return ws


def input_cell(ws, ref, value, fmt=NUM):
    c = ws[ref]
    c.value = value
    c.font = F_INPUT
    c.number_format = fmt
    c.alignment = RIGHT
    c.border = BOX
    return c


def calc_cell(ws, ref, formula, fmt=NUM):
    c = ws[ref]
    c.value = formula
    c.font = F_CALC
    c.number_format = fmt
    c.alignment = RIGHT
    return c


def header_row(ws, row, cols, texts, fill=True):
    for col, text in zip(cols, texts):
        c = ws[f"{col}{row}"]
        c.value = text
        c.font = F_HDRW if fill else F_BOLD
        c.alignment = CENTER if col != "A" else Alignment(horizontal="left")
        if fill:
            c.fill = FILL_NAVY


def howto(ws, row, lines):
    ws[f"A{row}"] = "HOW TO USE THIS TAB"
    ws[f"A{row}"].font = F_KICKER
    for i, line in enumerate(lines, start=1):
        c = ws[f"A{row + i}"]
        c.value = "• " + line
        c.font = F_MUTED
    src = ws[f"A{row + len(lines) + 2}"]
    src.value = "Source: [replace before presenting — source line is mandatory on every exhibit]"
    src.font = F_SOURCE


# ======================================================================================
wb = Workbook()
wb.remove(wb.active)

# ---- 0. READ ME ----------------------------------------------------------------------
ws = wb.create_sheet("READ ME")
ws.sheet_properties.tabColor = INK
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for col, w in (("B", 24), ("C", 60), ("D", 14), ("E", 30)):
    ws.column_dimensions[col].width = w

ws["B2"] = "CASE DESIGN SYSTEM"
ws["B2"].font = F_KICKER
ws["B3"] = "Excel chart templates — corporate family"
ws["B3"].font = Font(name=FONT, size=18, bold=True, color=INK)
ws["B4"] = ("Native Excel charts pre-styled to tokens/tokens.json and docs/CHART_RULES.md. "
            "Edit the blue input cells on any tab; the chart updates live. Paste charts into "
            "slides while numbers are still moving; once locked, hand the tab to the SVG builder "
            "for the pixel-perfect version (see 'Live data vs locked data' in CHART_RULES).")
ws["B4"].font = F_BODY
ws["B4"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[4].height = 48

rules = [
    ("Blue cells = inputs", "Type over them. Grey cells are helper formulas the chart needs — do not delete, they recalc."),
    ("Direct labels over legends", "Most tabs have data labels on and the value axis off (rule 2). Legends appear only where segments are too thin (stacked bars)."),
    ("Grey the field, colour the hero", "Comparison charts ship with one accent element. To move the hero, recolour the single bar/point in Excel — never add a second accent."),
    ("Action title lives on the slide", "Charts here carry no title. Write the conclusion as the slide headline ('Revenue grows to $300m by 2028'), not on the chart."),
    ("Semantic colours are reserved", "Green #2E8B6F = upside, red #C0473E = downside, amber #E0A13C = caution/reference. Never use them for neutral categories."),
    ("Source line always", "Every tab has a source placeholder. Replace it; bottom-left, italic, grey on the slide."),
    ("Units once", "State the unit in a top-left note ('Revenues in $m'), one precision per chart."),
]
r = 6
ws[f"B{r}"] = "House rules baked into these templates"
ws[f"B{r}"].font = F_BOLD
for name, desc in rules:
    r += 1
    ws[f"B{r}"] = name
    ws[f"B{r}"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
    ws[f"C{r}"] = desc
    ws[f"C{r}"].font = F_MUTED
    ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 26

r += 2
ws[f"B{r}"] = "Tab index"
ws[f"B{r}"].font = F_BOLD
index = [
    ("Waterfall",        "Bridge from a start total through deltas to an end total (ref04)."),
    ("Waterfall-H",      "Horizontal lever bridge: levers stack top-to-bottom (ref10, rule 18)."),
    ("Football Field",   "Valuation ranges by methodology; usually locked → best SVG candidate."),
    ("Scenario Bars",    "Bear / base / bull outcome bars with fixed semantic colours (ref07/12)."),
    ("Build-Up Table",   "ARPU × users = revenue bottom-up table, no chart (ref01, rule 15)."),
    ("Column-Hero",      "Time series columns, grey field + one accent hero year (rule 4)."),
    ("Bars-Ranking",     "Horizontal ranking bars, labels left, hero highlighted (ref11, rule 18)."),
    ("Stacked Bars",     "Segments over time; legend allowed because segments are thin (rule 1)."),
    ("Line",             "Indexed lines, hero accent vs grey field."),
    ("Combo Bar+Line",   "Bars on left axis, rate line on right axis (rules 27–28)."),
    ("Radar",            "Spokes normalised 0–100 before plotting (rules 23–26)."),
    ("Pie",              "Share-of-total, ≤5 slices, category+percent labels, no legend."),
    ("Dual-Axis Lines",  "Two units, two scales, right axis coloured like its series (ref16)."),
]
for name, desc in index:
    r += 1
    ws[f"B{r}"] = name
    ws[f"B{r}"].font = Font(name=FONT, size=10, bold=True, color=ACCENT)
    ws[f"C{r}"] = desc
    ws[f"C{r}"].font = F_MUTED

r += 2
ws[f"B{r}"] = "Palette (corporate)"
ws[f"B{r}"].font = F_BOLD
swatches = [
    ("ink", INK, "Headline text"), ("navy", NAVY, "Primary bars, headers"),
    ("accent", ACCENT, "The single hero"), ("blueMid", BLUEMID, "Secondary series"),
    ("blueLight", BLUELIGHT, "Tertiary series"), ("slate", SLATE, "Muted labels, source"),
    ("gridline", GRID, "Axes, faint gridlines, grey field bars"), ("panel", PANEL, "Panel background"),
    ("positive", POS, "Upside only"), ("negative", NEG, "Downside only"), ("caution", CAUT, "Reference lines only"),
]
for name, hexv, use in swatches:
    r += 1
    ws[f"B{r}"] = f"{name}  #{hexv}"
    ws[f"B{r}"].font = Font(name=FONT, size=10, color=INK if hexv in (GRID, PANEL, BLUELIGHT, CAUT) else WHITE, bold=True)
    ws[f"B{r}"].fill = PatternFill("solid", fgColor=hexv)
    ws[f"C{r}"] = use
    ws[f"C{r}"].font = F_MUTED


# ---- 1. Waterfall (vertical) ---------------------------------------------------------
ws = sheet(wb, "Waterfall", "Finance exhibit", "Waterfall / bridge — start → deltas → end",
           "Edit labels and blue deltas. Bases recompute; totals are formulas. Values in $m.")
header_row(ws, 5, "ABCDEFG", ["Step", "Delta", "Cum.", "base", "up", "down", "total"])
rows = [("FY24 revenue", 240, "start"), ("Volume growth", 45, "d"), ("Pricing", 30, "d"),
        ("FX headwind", -12, "d"), ("Churn", -18, "d"), ("FY28 revenue", None, "end")]
r0 = 6
for i, (label, val, kind) in enumerate(rows):
    r = r0 + i
    ws[f"A{r}"] = label
    ws[f"A{r}"].font = F_BOLD if kind != "d" else F_BODY
    if kind == "start":
        input_cell(ws, f"B{r}", val)
        calc_cell(ws, f"C{r}", f"=B{r}")
        calc_cell(ws, f"D{r}", 0)
        calc_cell(ws, f"E{r}", "=NA()")
        calc_cell(ws, f"F{r}", "=NA()")
        calc_cell(ws, f"G{r}", f"=B{r}")
    elif kind == "d":
        input_cell(ws, f"B{r}", val)
        calc_cell(ws, f"C{r}", f"=C{r-1}+B{r}")
        calc_cell(ws, f"D{r}", f"=IF(B{r}>=0,C{r-1},C{r})")
        calc_cell(ws, f"E{r}", f"=IF(B{r}>=0,B{r},NA())")
        calc_cell(ws, f"F{r}", f"=IF(B{r}<0,-B{r},NA())")
        calc_cell(ws, f"G{r}", "=NA()")
    else:
        calc_cell(ws, f"B{r}", f"=C{r-1}")
        ws[f"B{r}"].font = F_BOLD
        calc_cell(ws, f"C{r}", f"=B{r}")
        calc_cell(ws, f"D{r}", 0)
        calc_cell(ws, f"E{r}", "=NA()")
        calc_cell(ws, f"F{r}", "=NA()")
        calc_cell(ws, f"G{r}", f"=B{r}")
rN = r0 + len(rows) - 1

ch = BarChart()
ch.type, ch.grouping, ch.overlap, ch.gapWidth = "col", "stacked", 100, 45
data = Reference(ws, min_col=4, max_col=7, min_row=5, max_row=rN)
cats = Reference(ws, min_col=1, min_row=r0, max_row=rN)
ch.add_data(data, titles_from_data=True)
ch.set_categories(cats)
ch.legend = None
invisible(ch.series[0])
solid(ch.series[3], NAVY)   # total
solid(ch.series[1], POS)    # up
solid(ch.series[2], NEG)    # down
labels(ch.series[3], NUM, "ctr", WHITE)
labels(ch.series[1], NUM, "ctr", WHITE)
labels(ch.series[2], '"-"#,##0', "ctr", WHITE)
style_axes(ch, val_delete=True)
ch.width, ch.height = 17, 9.5
ws.add_chart(ch, "I4")
howto(ws, rN + 3, [
    "Add a step: insert a row inside the delta block, copy formulas C:G from the row above, extend the chart range.",
    "Green/red are semantic (rule 5); the endpoint bars stay navy. Do not recolour deltas.",
    "Labels sit inside segments (Excel limit); the SVG waterfall component floats them above — hand off once locked.",
])

# ---- 2. Waterfall-H (horizontal lever bridge, ref10) -----------------------------------
ws = sheet(wb, "Waterfall-H", "Finance exhibit", "Horizontal waterfall — cost/lever bridge",
           "Levers stack top-to-bottom from start to end (ref10, rule 18). Values in $/unit.")
header_row(ws, 5, "ABCDEFG", ["Lever", "Delta", "Cum.", "base", "up", "down", "total"])
rows = [("Cost today", 100, "start"), ("Automation", -18, "d"), ("Scale effects", -14, "d"),
        ("Energy mix", -9, "d"), ("Logistics", -5, "d"), ("Cost 2030", None, "end")]
r0 = 6
for i, (label, val, kind) in enumerate(rows):
    r = r0 + i
    ws[f"A{r}"] = label
    ws[f"A{r}"].font = F_BOLD if kind != "d" else F_BODY
    if kind == "start":
        input_cell(ws, f"B{r}", val)
        calc_cell(ws, f"C{r}", f"=B{r}")
        calc_cell(ws, f"D{r}", 0)
        calc_cell(ws, f"E{r}", "=NA()")
        calc_cell(ws, f"F{r}", "=NA()")
        calc_cell(ws, f"G{r}", f"=B{r}")
    elif kind == "d":
        input_cell(ws, f"B{r}", val)
        calc_cell(ws, f"C{r}", f"=C{r-1}+B{r}")
        calc_cell(ws, f"D{r}", f"=IF(B{r}>=0,C{r-1},C{r})")
        calc_cell(ws, f"E{r}", f"=IF(B{r}>=0,B{r},NA())")
        calc_cell(ws, f"F{r}", f"=IF(B{r}<0,-B{r},NA())")
        calc_cell(ws, f"G{r}", "=NA()")
    else:
        calc_cell(ws, f"B{r}", f"=C{r-1}")
        ws[f"B{r}"].font = F_BOLD
        calc_cell(ws, f"C{r}", f"=B{r}")
        calc_cell(ws, f"D{r}", 0)
        calc_cell(ws, f"E{r}", "=NA()")
        calc_cell(ws, f"F{r}", "=NA()")
        calc_cell(ws, f"G{r}", f"=B{r}")
rN = r0 + len(rows) - 1
ws["I17"] = "Delta chip:"
ws["I17"].font = F_MUTED
c = ws["J17"]
c.value = f"=(G{rN}-G{r0})/G{r0}"
c.number_format = PCT
c.font = Font(name=FONT, size=14, bold=True, color=WHITE)
c.fill = FILL_NAVY
c.alignment = CENTER

ch = BarChart()
ch.type, ch.grouping, ch.overlap, ch.gapWidth = "bar", "stacked", 100, 45
data = Reference(ws, min_col=4, max_col=7, min_row=5, max_row=rN)
cats = Reference(ws, min_col=1, min_row=r0, max_row=rN)
ch.add_data(data, titles_from_data=True)
ch.set_categories(cats)
ch.legend = None
ch.x_axis.scaling.orientation = "maxMin"   # first lever at the top
invisible(ch.series[0])
solid(ch.series[3], NAVY)
solid(ch.series[1], POS)
solid(ch.series[2], NEG)
labels(ch.series[3], NUM, "ctr", WHITE)
labels(ch.series[1], NUM, "ctr", WHITE)
labels(ch.series[2], '"-"#,##0', "ctr", WHITE)
style_axes(ch, val_delete=True)
ch.width, ch.height = 17, 9.5
ws.add_chart(ch, "I4")
howto(ws, rN + 3, [
    "Reads top-to-bottom: start total, then each lever connects to the running total, end total at the foot.",
    "The navy % chip (J17) is the bold delta callout from ref10 — paste it beside the chart on the slide.",
    "Annotation column explaining each lever goes on the slide to the right of the chart, not in the chart.",
])

# ---- 3. Football Field ----------------------------------------------------------------
ws = sheet(wb, "Football Field", "Finance exhibit", "Football field — valuation ranges by methodology",
           "Edit blue low/high per methodology and the current price. Values in $/share.")
header_row(ws, 5, "ABCDE", ["Methodology", "Low", "High", "base", "range"])
ffrows = [("52-week range", 18, 31), ("Analyst targets", 24, 38), ("Trading comps", 26, 40),
          ("Precedent transactions", 30, 46), ("DCF", 32, 50)]
r0 = 6
for i, (label, lo, hi) in enumerate(ffrows):
    r = r0 + i
    ws[f"A{r}"] = label
    ws[f"A{r}"].font = F_BODY
    input_cell(ws, f"B{r}", lo, USD)
    input_cell(ws, f"C{r}", hi, USD)
    calc_cell(ws, f"D{r}", f"=B{r}", USD)
    calc_cell(ws, f"E{r}", f"=C{r}-B{r}", USD)
rN = r0 + len(ffrows) - 1
ws[f"A{rN+2}"] = "Current share price"
ws[f"A{rN+2}"].font = F_BOLD
c = input_cell(ws, f"B{rN+2}", 28, USD)
c.fill = FILL_CAUTBG

ch = BarChart()
ch.type, ch.grouping, ch.overlap, ch.gapWidth = "bar", "stacked", 100, 70
data = Reference(ws, min_col=4, max_col=5, min_row=5, max_row=rN)
cats = Reference(ws, min_col=1, min_row=r0, max_row=rN)
ch.add_data(data, titles_from_data=True)
ch.set_categories(cats)
ch.legend = None
ch.x_axis.scaling.orientation = "maxMin"
invisible(ch.series[0])
solid(ch.series[1], BLUELIGHT)
point_color(ch.series[1], len(ffrows) - 1, NAVY)   # hero methodology: DCF
style_axes(ch, val_delete=False, val_grid=True, val_fmt=USD)
ch.width, ch.height = 17, 9.5
ws.add_chart(ch, "G4")
howto(ws, rN + 4, [
    "Value axis stays on with faint gridlines — the reader must read prices off it (rule 7 exception).",
    "Hero methodology (DCF) is navy; the field is light blue. Move the hero by recolouring a single bar.",
    "Excel cannot draw the amber current-price line natively: add a vertical line shape at $"
    " current price, or hand off to components/charts-finance/football-field once numbers lock.",
])

# ---- 4. Scenario Bars ------------------------------------------------------------------
ws = sheet(wb, "Scenario Bars", "Finance exhibit", "Scenario bars — bear / base / bull",
           "Edit blue price targets. Colours are fixed semantics: red bear, blue base, green bull.")
header_row(ws, 5, "AB", ["Scenario", "Price target"])
for i, (label, v) in enumerate([("Bear", 18), ("Base", 27), ("Bull", 36)]):
    r = 6 + i
    ws[f"A{r}"] = label
    ws[f"A{r}"].font = F_BODY
    input_cell(ws, f"B{r}", v, USD)
ws["A10"] = "Current price"
ws["A10"].font = F_BOLD
input_cell(ws, "B10", 22, USD).fill = FILL_CAUTBG

ch = BarChart()
ch.type, ch.gapWidth = "col", 80
data = Reference(ws, min_col=2, min_row=5, max_row=8)
cats = Reference(ws, min_col=1, min_row=6, max_row=8)
ch.add_data(data, titles_from_data=True)
ch.set_categories(cats)
ch.legend = None
s = ch.series[0]
solid(s, ACCENT)
point_color(s, 0, NEG)
point_color(s, 1, ACCENT)
point_color(s, 2, POS)
labels(s, USD, "outEnd", INK, 1100)
style_axes(ch, val_delete=True)
ch.width, ch.height = 13, 9
ws.add_chart(ch, "D4")
howto(ws, 13, [
    "Bear/base/bull colour mapping is fixed in the token aliases — never repurpose it.",
    "Add the upside vs current price as a label or oval CAGR-style chip on the slide (ref08).",
])

# ---- 5. Build-Up Table -----------------------------------------------------------------
ws = sheet(wb, "Build-Up Table", "Finance exhibit", "Bottom-up build-up: users × ARPU = revenue",
           "Edit blue driver rows; revenue and growth are formulas. Show the arithmetic (rule 15).")
years = ["FY24", "FY25", "FY26", "FY27", "FY28"]
header_row(ws, 5, "ABCDEF", ["$m unless noted"] + years, fill=False)
for col in "BCDEF":
    ws[f"{col}5"].border = RULE_B
ws["A5"].border = RULE_B
drivers = [("Users (m)", [10.0, 11.5, 13.2, 15.0, 16.8], NUM1),
           ("× ARPU ($)", [24.0, 25.2, 26.5, 27.8, 29.2], NUM1)]
r = 6
for label, vals, fmt in drivers:
    ws[f"A{r}"] = label
    ws[f"A{r}"].font = F_BODY
    for j, v in enumerate(vals):
        input_cell(ws, f"{get_column_letter(2+j)}{r}", v, fmt)
    r += 1
ws[f"A{r}"] = '="= Revenue ($m)"'   # leading "=" as literal text would parse as a formula
ws[f"A{r}"].font = F_BOLD
for j in range(5):
    col = get_column_letter(2 + j)
    c = calc_cell(ws, f"{col}{r}", f"={col}6*{col}7", NUM)
    c.font = F_BOLD
    c.border = RULE_T
ws[f"A{r}"].border = RULE_T
rev_row = r
r += 1
ws[f"A{r}"] = "YoY growth"
ws[f"A{r}"].font = F_BODY
ws[f"B{r}"] = "—"
ws[f"B{r}"].font = F_CALC
ws[f"B{r}"].alignment = RIGHT
for j in range(1, 5):
    col, prev = get_column_letter(2 + j), get_column_letter(1 + j)
    c = calc_cell(ws, f"{col}{r}", f"={col}{rev_row}/{prev}{rev_row}-1", PCT1)
    c.fill = FILL_POSBG
    c.font = Font(name=FONT, size=10, color=POS, bold=True)
r += 2
ws[f"A{r}"] = "Revenue CAGR FY24–28"
ws[f"A{r}"].font = F_BOLD
c = calc_cell(ws, f"B{r}", f"=(F{rev_row}/B{rev_row})^(1/4)-1", PCT1)
c.font = Font(name=FONT, size=12, bold=True, color=INK)
howto(ws, r + 2, [
    "Tables use horizontal rules only (rule 20): rule under the header, rule above the total row.",
    "Right-align numbers, one precision per row (rules 11–12). Unit stated once in the corner cell.",
    "Close the slide with a navy takeaway strip stating the conclusion (ref01, rule 17).",
])

# ---- 6. Column-Hero --------------------------------------------------------------------
ws = sheet(wb, "Column-Hero", "Core chart", "Columns over time — grey field, one hero",
           "Edit blue values. The hero year is accent; every other bar stays grey (rule 4).", tab=ACCENT)
header_row(ws, 5, "AB", ["Year", "Revenue ($m)"])
vals = [("FY22", 120), ("FY23", 138), ("FY24", 161), ("FY25", 190), ("FY26", 224), ("FY27", 262), ("FY28", 300)]
for i, (y, v) in enumerate(vals):
    r = 6 + i
    ws[f"A{r}"] = y
    ws[f"A{r}"].font = F_BODY
    input_cell(ws, f"B{r}", v)
rN = 5 + len(vals)

ch = BarChart()
ch.type, ch.gapWidth = "col", 55
data = Reference(ws, min_col=2, min_row=5, max_row=rN)
cats = Reference(ws, min_col=1, min_row=6, max_row=rN)
ch.add_data(data, titles_from_data=True)
ch.set_categories(cats)
ch.legend = None
s = ch.series[0]
solid(s, GRID)
point_color(s, len(vals) - 1, ACCENT)
labels(s, NUM, "outEnd", INK)
style_axes(ch, val_delete=True)
ch.width, ch.height = 16, 9
ws.add_chart(ch, "D4")
howto(ws, rN + 3, [
    "Labels on, axis off (rule 2). One accent max (rule 6) — move it, don't add a second.",
    "The hero's label can be bolded/enlarged in Excel to be the loudest number (rule 3).",
])

# ---- 7. Bars-Ranking -------------------------------------------------------------------
ws = sheet(wb, "Bars-Ranking", "Core chart", "Horizontal ranking bars",
           "Ranking/scenario lists read sideways with labels on the left (rule 18, ref11).", tab=ACCENT)
header_row(ws, 5, "AB", ["Player", "Share (%)"])
rank = [("Peer A", 24), ("Us", 21), ("Peer B", 18), ("Peer C", 12), ("Peer D", 8)]
for i, (label, v) in enumerate(rank):
    r = 6 + i
    ws[f"A{r}"] = label
    ws[f"A{r}"].font = F_BODY
    input_cell(ws, f"B{r}", v)
rN = 5 + len(rank)

ch = BarChart()
ch.type, ch.gapWidth = "bar", 55
data = Reference(ws, min_col=2, min_row=5, max_row=rN)
cats = Reference(ws, min_col=1, min_row=6, max_row=rN)
ch.add_data(data, titles_from_data=True)
ch.set_categories(cats)
ch.legend = None
ch.x_axis.scaling.orientation = "maxMin"   # keep input order, largest on top
s = ch.series[0]
solid(s, GRID)
point_color(s, 1, ACCENT)                  # hero: "Us"
labels(s, '0"%"', "outEnd", INK)
style_axes(ch, val_delete=True)
ch.width, ch.height = 15, 8.5
ws.add_chart(ch, "D4")
howto(ws, rN + 3, [
    "Sort the input rows descending yourself — the chart plots in row order, first row on top.",
    "The hero bar is wired to row 2 ('Us'). If the ranking changes, move the accent to the new hero row.",
])

# ---- 8. Stacked Bars -------------------------------------------------------------------
ws = sheet(wb, "Stacked Bars", "Core chart", "Stacked bars by segment",
           "Legend is the fallback here because segments are thin (rule 1). Values in $m.", tab=ACCENT)
header_row(ws, 5, "ABCDE", ["Segment"] + ["FY25", "FY26", "FY27", "FY28"])
seg = [("Core", [140, 152, 163, 172]), ("Growth", [34, 52, 74, 96]), ("New bets", [8, 16, 27, 40])]
for i, (label, vals) in enumerate(seg):
    r = 6 + i
    ws[f"A{r}"] = label
    ws[f"A{r}"].font = F_BODY
    for j, v in enumerate(vals):
        input_cell(ws, f"{get_column_letter(2+j)}{r}", v)
rN = 8

ch = BarChart()
ch.type, ch.grouping, ch.overlap, ch.gapWidth = "col", "stacked", 100, 60
data = Reference(ws, min_col=1, max_col=5, min_row=6, max_row=rN)
cats = Reference(ws, min_col=2, max_col=5, min_row=5)
ch.add_data(data, titles_from_data=True, from_rows=True)
ch.set_categories(cats)
for s, colr in zip(ch.series, (NAVY, BLUEMID, BLUELIGHT)):
    solid(s, colr)
labels(ch.series[0], NUM, "ctr", WHITE)
labels(ch.series[1], NUM, "ctr", WHITE)
labels(ch.series[2], NUM, "ctr", INK)
ch.legend.position = "t"
ch.legend.txPr = rich(1000, SLATE)
style_axes(ch, val_delete=True)
ch.width, ch.height = 15, 9
ws.add_chart(ch, "G4")
howto(ws, rN + 3, [
    "Series palette runs navy → blueMid → blueLight, biggest segment darkest and at the base.",
    "Label the largest segments directly even with a legend on (rule 1).",
    "Summary total ovals or a CAGR arrow across bar tops are slide annotations (ref09/13), added after paste.",
])

# ---- 9. Line ---------------------------------------------------------------------------
ws = sheet(wb, "Line", "Core chart", "Lines over time — hero vs field",
           "Indexed series (100 = 2019). Hero is accent and thicker; the field is grey.", tab=ACCENT)
header_row(ws, 5, "ABCDEFGH", ["Series"] + ["2019", "2020", "2021", "2022", "2023", "2024", "2025"])
lines = [("Us", [100, 104, 118, 131, 152, 178, 205]), ("Peer median", [100, 101, 108, 112, 121, 128, 134])]
for i, (label, vals) in enumerate(lines):
    r = 6 + i
    ws[f"A{r}"] = label
    ws[f"A{r}"].font = F_BODY
    for j, v in enumerate(vals):
        input_cell(ws, f"{get_column_letter(2+j)}{r}", v)

ch = LineChart()
data = Reference(ws, min_col=1, max_col=8, min_row=6, max_row=7)
cats = Reference(ws, min_col=2, max_col=8, min_row=5)
ch.add_data(data, titles_from_data=True, from_rows=True)
ch.set_categories(cats)
ch.legend = None
line_style(ch.series[0], ACCENT, 2.5)
line_style(ch.series[1], SLATE, 1.5)
style_axes(ch, val_delete=False, val_grid=True, val_fmt=NUM)
ch.width, ch.height = 16, 9
ws.add_chart(ch, "C10")
howto(ws, 22, [
    "Direct side labels at line ends beat a legend (rule 1): add two text boxes ('Us', 'Peer median') after pasting.",
    "Keep the axis: precise values matter on indexed lines, so faint gridlines stay (rule 7).",
])

# ---- 10. Combo Bar+Line ----------------------------------------------------------------
ws = sheet(wb, "Combo Bar+Line", "Core chart", "Bars + rate line, two axes by unit",
           "One scale per unit (rule 27): $m bars own the left axis, margin % owns the right.", tab=ACCENT)
header_row(ws, 5, "ABCDEF", ["Series"] + ["FY24", "FY25", "FY26", "FY27", "FY28"])
ws["A6"] = "Revenue ($m)"
ws["A6"].font = F_BODY
ws["A7"] = "EBITDA margin"
ws["A7"].font = F_BODY
for j, v in enumerate([182, 204, 238, 268, 300]):
    input_cell(ws, f"{get_column_letter(2+j)}6", v)
for j, v in enumerate([0.18, 0.20, 0.22, 0.24, 0.26]):
    input_cell(ws, f"{get_column_letter(2+j)}7", v, PCT)

bar = BarChart()
bar.type, bar.gapWidth = "col", 60
bar.add_data(Reference(ws, min_col=1, max_col=6, min_row=6, max_row=6), titles_from_data=True, from_rows=True)
bar.set_categories(Reference(ws, min_col=2, max_col=6, min_row=5))
solid(bar.series[0], NAVY)
style_axes(bar, val_delete=False, val_grid=False, val_fmt=NUM)
bar.y_axis.title = "$m"
bar.y_axis.title.tx.rich.p[0].pPr = ParagraphProperties(
    defRPr=CharacterProperties(latin=DrawFont(typeface=FONT), sz=1000, solidFill=SLATE))

ln = LineChart()
ln.add_data(Reference(ws, min_col=1, max_col=6, min_row=7, max_row=7), titles_from_data=True, from_rows=True)
line_style(ln.series[0], ACCENT, 2.5)
labels(ln.series[0], PCT, "t", ACCENT)
ln.y_axis.axId = 200
ln.y_axis.crosses = "max"
ln.y_axis.majorGridlines = None
ln.y_axis.majorTickMark = "none"
ln.y_axis.number_format = PCT
ln.y_axis.txPr = rich(1100, ACCENT)       # right axis styled in the line's colour (rule 27)
ln.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
bar += ln
bar.legend = None
bar.width, bar.height = 16, 9.5
ws.add_chart(bar, "C10")
howto(ws, 23, [
    "Two axes maximum; group any extra series by unit onto an existing axis (rule 27).",
    "Right axis text is accent — same colour as the line — so ownership is unambiguous.",
    "Axis maxima: round up to a clean interval so gridlines land on round numbers (rule 28).",
])

# ---- 11. Radar -------------------------------------------------------------------------
ws = sheet(wb, "Radar", "Specialty", "Radar — normalise every axis to 0–100 first",
           "Enter raw metrics + the 100-score benchmark; spokes plot the normalised block (rule 23).", tab=BLUEMID)
header_row(ws, 5, "ABCD", ["Metric (raw)", "Us", "Peer", "Scale max"])
radar_rows = [("Revenue growth (%)", 24, 15, 30), ("EBITDA margin (%)", 22, 26, 35),
              ("Market share (%)", 21, 24, 30), ("NPS (pts)", 46, 31, 60), ("Retention (%)", 88, 80, 100)]
r0 = 6
for i, (m, us, peer, mx) in enumerate(radar_rows):
    r = r0 + i
    ws[f"A{r}"] = m
    ws[f"A{r}"].font = F_BODY
    input_cell(ws, f"B{r}", us)
    input_cell(ws, f"C{r}", peer)
    input_cell(ws, f"D{r}", mx)
rN = r0 + len(radar_rows) - 1
hr = rN + 2
header_row(ws, hr, "ABC", ["Normalised 0–100", "Us", "Peer"])
for i in range(len(radar_rows)):
    r_src, r = r0 + i, hr + 1 + i
    calc_cell(ws, f"A{r}", f"=A{r_src}", "General")
    ws[f"A{r}"].alignment = Alignment(horizontal="left")
    calc_cell(ws, f"B{r}", f"=ROUND(B{r_src}/D{r_src}*100,0)")
    calc_cell(ws, f"C{r}", f"=ROUND(C{r_src}/D{r_src}*100,0)")
nN = hr + len(radar_rows)

ch = RadarChart()
ch.type = "standard"
data = Reference(ws, min_col=2, max_col=3, min_row=hr, max_row=nN)
cats = Reference(ws, min_col=1, min_row=hr + 1, max_row=nN)
ch.add_data(data, titles_from_data=True)
ch.set_categories(cats)
line_style(ch.series[0], ACCENT, 2)
line_style(ch.series[1], GRID, 1.5)
ch.y_axis.scaling.min, ch.y_axis.scaling.max = 0, 100
ch.y_axis.majorUnit = 25                    # 4 rings max (rule 26)
ch.y_axis.majorGridlines = faint_gridlines()
ch.y_axis.txPr = rich(900, SLATE)
ch.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
ch.x_axis.txPr = rich(1000, SLATE)
ch.legend.position = "t"
ch.legend.txPr = rich(1000, SLATE)
ch.width, ch.height = 13, 11
ws.add_chart(ch, "F4")
howto(ws, nN + 2, [
    "Never plot raw units on spokes — the 0–100 block is what the chart reads (rule 23).",
    "Put the most important metric in the first row: Excel places it at 12 o'clock (rule 24).",
    "Keep the scoring table with the chart so the exhibit can be re-derived.",
])

# ---- 12. Pie ---------------------------------------------------------------------------
ws = sheet(wb, "Pie", "Specialty", "Share of total — ≤5 slices, labelled directly",
           "Category + percent labels, no legend (rule 1). Order slices largest-first.", tab=BLUEMID)
header_row(ws, 5, "AB", ["Slice", "Value ($m)"])
slices = [("North America", 96), ("Europe", 68), ("APAC", 45), ("Rest of world", 21)]
for i, (label, v) in enumerate(slices):
    r = 6 + i
    ws[f"A{r}"] = label
    ws[f"A{r}"].font = F_BODY
    input_cell(ws, f"B{r}", v)
rN = 5 + len(slices)

ch = PieChart()
data = Reference(ws, min_col=2, min_row=5, max_row=rN)
cats = Reference(ws, min_col=1, min_row=6, max_row=rN)
ch.add_data(data, titles_from_data=True)
ch.set_categories(cats)
ch.legend = None
s = ch.series[0]
for i, colr in enumerate((NAVY, BLUEMID, BLUELIGHT, GRID)):
    point_color(s, i, colr)
dl = DataLabelList(showVal=False, showSerName=False, showCatName=True, showLegendKey=False,
                   showPercent=True, showBubbleSize=False)
dl.dLblPos = "outEnd"
dl.txPr = rich(1000, INK)
ch.dLbls = dl
ch.firstSliceAng = 0
ch.width, ch.height = 12, 10
ws.add_chart(ch, "D4")
howto(ws, rN + 3, [
    "Four neutral blues/greys — colour is not meaning here, so no semantic colours (rule 5).",
    "More than 5 categories? Merge the tail into 'Other' or switch to the ranking-bars tab.",
    "For the leader-line look of ref05, drag labels outward in Excel; leader lines appear automatically.",
])

# ---- 13. Dual-Axis Lines ---------------------------------------------------------------
ws = sheet(wb, "Dual-Axis Lines", "Specialty", "Two units, two scales (ref16)",
           "Left axis: $m (navy). Right axis: count (accent) — axis coloured like its series.", tab=BLUEMID)
header_row(ws, 5, "ABCDEFGH", ["Series"] + ["2019", "2020", "2021", "2022", "2023", "2024", "2025"])
ws["A6"] = "Investment ($m)"
ws["A6"].font = F_BODY
ws["A7"] = "Deal count"
ws["A7"].font = F_BODY
for j, v in enumerate([310, 280, 520, 640, 590, 830, 1050]):
    input_cell(ws, f"{get_column_letter(2+j)}6", v)
for j, v in enumerate([14, 12, 21, 26, 24, 31, 38]):
    input_cell(ws, f"{get_column_letter(2+j)}7", v)

l1 = LineChart()
l1.add_data(Reference(ws, min_col=1, max_col=8, min_row=6, max_row=6), titles_from_data=True, from_rows=True)
l1.set_categories(Reference(ws, min_col=2, max_col=8, min_row=5))
line_style(l1.series[0], NAVY, 2.5)
style_axes(l1, val_delete=False, val_grid=False, val_fmt=NUM)
l1.y_axis.txPr = rich(1100, NAVY)

l2 = LineChart()
l2.add_data(Reference(ws, min_col=1, max_col=8, min_row=7, max_row=7), titles_from_data=True, from_rows=True)
line_style(l2.series[0], ACCENT, 2.5)
l2.y_axis.axId = 210
l2.y_axis.crosses = "max"
l2.y_axis.majorGridlines = None
l2.y_axis.majorTickMark = "none"
l2.y_axis.txPr = rich(1100, ACCENT)
l2.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
l1 += l2
l1.legend = None
l1.width, l1.height = 16, 9.5
ws.add_chart(l1, "C10")
howto(ws, 23, [
    "Each axis takes the colour of the series that owns it — the reader never guesses which scale is whose.",
    "Label each line directly at its right end with a text box after pasting (rule 1).",
    "Use this only when both series genuinely share the time axis; otherwise make two small charts.",
])

wb.save("templates/chart-templates.xlsx")
print("Wrote templates/chart-templates.xlsx with sheets:", wb.sheetnames)
